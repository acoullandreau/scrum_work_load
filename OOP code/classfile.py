import csv
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from string import ascii_uppercase


def map_columns(array, name):
    for i in range(0, len(array)):
        if array[i] == name:
            return i
            break


class JiraToolBox:

    def __init__(self, config, output_type):
        self.config = config
        self.output_type = output_type
        self.row_min = self.config['row_min']
        self.issue_col = self.config['issue_col']
        self.status_col = self.config['status_col']
        self.priority_col = self.config['priority_col']
        self.story_points_col = self.config['points_col']
        self.first_assignee_col = self.config['first_assignee_col']
        self.assignee_row = self.config['assignee_row']
        self.csv_issues_id_dict = {}
        self.csv_issues_key_dict = {}
        self.csv_structure = {}
        self.xls_issues_list = []
        self.source_file = None
        self.assignees = {}
        self.assignees_story_remaining_estimate = {}
        self.assignees_processed = False


    def parseJiraIssues(self, csv_filename):
        with open(csv_filename) as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            data = list(readCSV)
            row_id = 1
            # maps each header to a column number based on the label
            for row in data:
                if row_id == 1:
                    self.csv_structure['header'] = row
                    self.csv_structure['Issue key'] = map_columns(row, 'Issue key')
                    self.csv_structure['Type'] = map_columns(row, 'Issue Type')
                    self.csv_structure['Issue id'] = map_columns(row, 'Issue id')
                    self.csv_structure['Parent id'] = map_columns(row, 'Parent id')
                    self.csv_structure['Summary'] = map_columns(row, 'Summary')
                    self.csv_structure['Assignee'] = map_columns(row, 'Assignee')
                    self.csv_structure['Priority'] = map_columns(row, 'Priority')
                    self.csv_structure['Status'] = map_columns(row, 'Status')
                    self.csv_structure['Story Points'] = map_columns(row, 'Custom field (Story Points)')
                    self.csv_structure['Remaining Estimate'] = map_columns(row, 'Î£ Remaining Estimate')
                    row_id += 1
                else:
            #populates the csv issue key dict with the key of an issue and its Issue instance
                    issue = Issue(row, self.csv_structure)
                    issue_key_col = self.csv_structure['Issue key']
                    self.csv_issues_key_dict[row[issue_key_col]] = issue


    def processJiraIssues(self):
        for key in self.csv_issues_key_dict.keys():
            #map each issue id to an issue key
            issue_id = self.csv_issues_key_dict[key].get_id()
            self.csv_issues_id_dict[issue_id] = key

        # maps each issue to its list of children instances and each issue to the instance of its parent
        for key in self.csv_issues_key_dict.keys():
            parent_id = self.csv_issues_key_dict[key].get_parent_id()
            if parent_id != '':
                parent_key = self.csv_issues_id_dict[parent_id]
                self.csv_issues_key_dict[key].set_parent(self.csv_issues_key_dict[parent_key])
                self.csv_issues_key_dict[parent_key].add_child(self.csv_issues_key_dict[key])

    def parseExistingFile(self, filename):
        self.source_file = load_workbook(filename)
        sheet = self.source_file.active
        row_max = sheet.max_row

        #lists the issues already present in the input file
        for i in range(self.row_min, row_max):
            issue_key = sheet.cell(row=i, column=self.issue_col).value
            if issue_key not in self.xls_issues_list:
                self.xls_issues_list.append(issue_key)


    def processIssueOwnership(self):
        # builds a dictionary of all assignees of a given standard issue
        if self.assignees_processed == False:
            self.processAssignees()
        for key in self.csv_issues_key_dict.keys():
            issue_assignee = self.csv_issues_key_dict[key].get_assignee()
            assignee_instance = self.assignees[issue_assignee]
            self.csv_issues_key_dict[key].assignee = assignee_instance

    def processAssignees(self):
        #builds a dictionary of all assignees listed in the csv file
        for key in self.csv_issues_key_dict.keys():
            issue_assignee = self.csv_issues_key_dict[key].get_assignee()
            if issue_assignee not in self.assignees.keys():
                self.assignees[issue_assignee] = Assignee(issue_assignee)
            self.assignees[issue_assignee].add_issue(self.csv_issues_key_dict[key])
        self.assignees_processed = True

    def processAssigneesRemainingEstimate(self):
        #builds a dictionary of issues key, which value is a dictionary of assignees with their remaining estimate for this issue
        for key in self.csv_issues_key_dict.keys():
            if self.csv_issues_key_dict[key].is_standard() == True:
                if self.csv_issues_key_dict[key].get_children() != []:
                    for item in self.csv_issues_key_dict[key].get_children():
                        item_key = item.get_key()
                        self.calculateRemainingTime(item_key,self.csv_issues_key_dict[item_key].get_assignee().get_name())
                else:
                    self.calculateRemainingTime(key,self.csv_issues_key_dict[key].get_assignee().get_name())

    def calculateRemainingTime(self, issue, assignee):
        if self.csv_issues_key_dict[issue].get_remaining_time() != '':
            if issue not in self.assignees_story_remaining_estimate.keys():
                if self.csv_issues_key_dict[issue].is_standard() == True:
                    self.assignees_story_remaining_estimate[issue] = {}
                    issue_assignee_remaining_time = int(self.csv_issues_key_dict[issue].get_remaining_time()) / 3600
                    self.assignees_story_remaining_estimate[issue][self.csv_issues_key_dict[issue].get_assignee().get_name()] = issue_assignee_remaining_time
                else:
                    parent_key = self.csv_issues_id_dict[self.csv_issues_key_dict[issue].get_parent_id()]
                    issue_assignee_remaining_time = 0
                    if parent_key not in self.assignees_story_remaining_estimate.keys():
                        self.assignees_story_remaining_estimate[parent_key] = {}
                    else:
                        if assignee in self.assignees_story_remaining_estimate[parent_key].keys():
                            issue_assignee_remaining_time = self.assignees_story_remaining_estimate[parent_key][self.csv_issues_key_dict[issue].get_assignee().get_name()]
                    issue_assignee_remaining_time = issue_assignee_remaining_time + int(self.csv_issues_key_dict[issue].get_remaining_time()) / 3600
                    self.assignees_story_remaining_estimate[parent_key][self.csv_issues_key_dict[issue].get_assignee().get_name()] = issue_assignee_remaining_time
            else:
                issue_assignee = self.csv_issues_key_dict[issue].get_assignee()
                if issue_assignee in self.assignees_story_remaining_estimate[issue].keys:
                    issue_assignee_remaining_time = self.assignees_story_remaining_estimate[issue][assignee]
                    issue_assignee_remaining_time = issue_assignee_remaining_time + int(self.csv_issues_key_dict[issue].get_remaining_time()) / 3600
                    self.assignees_story_remaining_estimate[issue][self.csv_issues_key_dict[issue].get_assignee().get_name()] = issue_assignee_remaining_time
                else:
                    self.assignees_story_remaining_estimate[issue][self.csv_issues_key_dict[issue].get_assignee().get_name()] = int(self.csv_issues_key_dict[issue].get_remaining_time()) / 3600
        else:
            if issue not in self.assignees_story_remaining_estimate.keys():
                if self.csv_issues_key_dict[issue].is_standard() == True:
                    self.assignees_story_remaining_estimate[issue] = {}
                    self.assignees_story_remaining_estimate[issue][self.csv_issues_key_dict[issue].get_assignee().get_name()] = 0
                else:
                    parent_key = self.csv_issues_id_dict[self.csv_issues_key_dict[issue].get_parent_id()]
                    if parent_key not in self.assignees_story_remaining_estimate.keys():
                        self.assignees_story_remaining_estimate[parent_key] = {}
                    self.assignees_story_remaining_estimate[parent_key][self.csv_issues_key_dict[issue].get_assignee().get_name()] = 0
            else:
                issue_assignee = self.csv_issues_key_dict[issue].get_assignee()
                if issue_assignee not in self.assignees_story_remaining_estimate[issue].keys:
                    self.assignees_story_remaining_estimate[issue][self.csv_issues_key_dict[issue].get_assignee().get_name()] = 0


    def processNewIssues(self):
        new_issues = []
        #add new lines if not in the file, updates the line otherwise
        for key in self.csv_issues_key_dict.keys():
            test_issue = self.csv_issues_key_dict[key].is_standard()
            if test_issue == True:
                if key not in self.xls_issues_list:
                    new_issues.append(key)
        return new_issues

    def writeIssueDetails(self, row, issue):
        sheet = self.source_file.active
        col_max = sheet.max_column
        # writes summary
        sheet.cell(row=row, column=col_max).value = self.csv_issues_key_dict[issue].get_summary()
        if self.output_type == 'matrix':
            # writes priority
            sheet.cell(row=row, column=self.priority_col).value = self.csv_issues_key_dict[issue].get_priority()
            # writes status
            sheet.cell(row=row, column=self.status_col).value = self.csv_issues_key_dict[issue].get_status()
            # writes story points
            if self.csv_issues_key_dict[issue].get_story_points() != '':
                sheet.cell(row=row, column=self.story_points_col).value = float(self.csv_issues_key_dict[issue].get_story_points())
            # processes assignee
            for j in range(self.first_assignee_col, col_max):
                assignee_jira_name = self.config[sheet.cell(row=self.assignee_row, column=j).value]
                for assigned_issue in self.assignees[assignee_jira_name].get_issue_list():
                    if assigned_issue.get_key() == issue:
                        sheet.cell(row=row, column=j).value = 'x'
        elif self.output_type == 'load':
            # processes remaining time
            for j in range(self.first_assignee_col, col_max):
                assignee_jira_name = self.config[sheet.cell(row=self.assignee_row, column=j).value]
                if assignee_jira_name in self.assignees_story_remaining_estimate[issue].keys():
                    sheet.cell(row=row, column=j).value = self.assignees_story_remaining_estimate[issue][assignee_jira_name]
                    sheet.cell(row=row, column=j).number_format = '#,##0.00'

    def addNewIssues(self, new_issues):
        sheet = self.source_file.active
        row_max = sheet.max_row
        index = max(row_max, self.row_min)
        #adds new rows for new issues
        for item in new_issues:
            issue_status = self.csv_issues_key_dict[item].get_status()
            if issue_status == 'Done':
                sheet.insert_rows(self.row_min + 1)
                # writes the issue key
                sheet.cell(row=self.row_min + 1, column=self.issue_col).value = item
                self.writeIssueDetails(self.row_min + 1, item)
            else:
                # writes the issue key
                sheet.cell(row=index, column=self.issue_col).value = item
                self.writeIssueDetails(index, item)
            index += 1


    def updateExistingIssues(self):
        sheet = self.source_file.active
        row_max = sheet.max_row
        col_max = sheet.max_column
        #updates the lines already existing in the file (without changing the position of the line)
        for key in self.csv_issues_key_dict.keys():
            test_issue = self.csv_issues_key_dict[key].is_standard()
            if test_issue == True:
                if key in self.xls_issues_list:
                    for i in range(self.row_min, row_max):
                        current_issue = sheet.cell(row=i, column=self.issue_col).value
                        if key == current_issue:
                            self.writeIssueDetails(i, key)


    def updateTotalsPerAssignee(self):
        #updates the total of the values written for each assignee (ex: remaining time)
        sheet = self.source_file.active
        row_max = sheet.max_row
        col_max = sheet.max_column
        total_row = 1

        for i in range(1, self.row_min):
            total = sheet.cell(row=i, column=self.issue_col).value
            if total == 'Allocated hours in sprint':
                total_row = i
                break

        for j in range(self.first_assignee_col, col_max):
            total = sheet.cell(row=total_row, column=j)
            top_cell_row = str(sheet.cell(row=self.row_min, column=j).row)
            top_cell_col = sheet.cell(row=self.row_min, column=j).column
            top_cell_col = str(ascii_uppercase[top_cell_col - 1])
            total_cell_row = str(sheet.cell(row=row_max, column=j).row)
            total_cell_col = sheet.cell(row=row_max, column=j).column
            total_cell_col = str(ascii_uppercase[total_cell_col - 1])
            total.value = '= SUM(' + top_cell_col + top_cell_row + ':' + total_cell_col + total_cell_row + ')'

    def updateTotalsPerPredictionSection(self):
        #parses the limits of each section, and computes the total sum of a given metric for this section (ex: story points)

        sheet = self.source_file.active
        row_max = sheet.max_row
        total_row = {}

        for i in range(self.row_min, row_max):
            category = sheet.cell(row=i, column=1).value
            if category != None and category != '':
                total_row[category] = []
                total_row[category].append(i)

        for key in total_row.keys():
            for i in range(total_row[key][0], row_max):
                total = sheet.cell(row=i, column=4).value
                if total == 'Total':
                    total_row[key].append(i)
                    break

        for key in total_row.keys():
            if key != None and key != 'To be sorted':
                top_cell = 'E' + str(total_row[key][0])
                total_cell = 'E' + str(total_row[key][1] - 1)
                total = sheet.cell(row=total_row[key][1], column=5)
                total.value = '= SUM(' + top_cell + ':' + total_cell + ')'

    def highlightRemovedIssues(self):
        #highlights the lines that were present in the source file and are not present in the input csv file

        sheet = self.source_file.active
        row_max = sheet.max_row
        col_max = sheet.max_column

        cell_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid")
        for i in range(self.row_min, row_max):
            story_key = sheet.cell(row=i, column=self.issue_col)
            if story_key.value != None:
                if story_key.value not in self.csv_issues_key_dict.keys():
                    for j in range(1, col_max +1):
                        sheet.cell(row=i, column=j).fill = cell_fill


    def writeUpdatedFile(self, output_filename):
        self.source_file.save(output_filename)



class Issue:

    def __init__(self, row, structure):
        self.key = row[structure['Issue key']]
        self.type = row[structure['Type']]
        self.id = row[structure['Issue id']]
        self.parent_id = row[structure['Parent id']]
        self.summary = row[structure['Summary']]
        self.assignee = row[structure['Assignee']]
        self.priority = row[structure['Priority']]
        self.status = row[structure['Status']]
        self.points = row[structure['Story Points']]
        self.remaining_time = row[structure['Remaining Estimate']]
        self.children = []

    def get_key(self):
        return self.key

    def get_type(self):
        return self.type

    def get_id(self):
        return self.id

    def get_parent_id(self):
        return self.parent_id

    def get_summary(self):
        return self.summary

    def get_assignee(self):
        return self.assignee

    def get_priority(self):
        return self.priority

    def get_status(self):
        return self.status

    def get_story_points(self):
        return self.points

    def get_remaining_time(self):
        return self.remaining_time

    def get_children(self):
        return self.children

    def is_standard(self):
        if self.type != 'Sub-task':
            return True
        else:
            return False

    def set_parent(self, issue):
        self.parent = issue

    def add_child(self, issue):
        self.children.append(issue)



class Assignee:

    def __init__(self, name):
        self.name = name
        self.issue_list = []

    def get_name(self):
        return self.name

    def add_issue(self, issue):
        self.issue_list.append(issue)

    def get_issue_list(self):
        return self.issue_list
