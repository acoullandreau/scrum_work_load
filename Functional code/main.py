import csv
import json
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from string import ascii_uppercase

dic_id_map = {}
dic_parent_map = {}
dic_assignee = {}
dic_story_map = {}
new_stories = []

with open('conf.json', encoding='utf-8') as config_file:
    conf_data = json.load(config_file)

def map_columns(array, name):
    for i in range(0, len(array)):
        if array[i] == name:
            return i
            break

def list_stories(key, id, parent):
    if parent == '':
        map_parent(key, id)
    else:
        story_key = dic_id_map[parent]
        map_parent(story_key, id)

def map_parent(key, id):
    if key not in dic_parent_map.keys():
        dic_parent_map[key] = []
    dic_parent_map[key].append(id)

def write_assignee(row_index,issue_key):
    for j in range(2, col_max - 5):
        assignee = sheet.cell(row=6, column=j).value
        assignee = conf_data[assignee]
        is_assignee = test_assignee(assignee, issue_key)
        if is_assignee == True:
            remain_time = remaining_time(assignee, issue_key)
            sheet.cell(row=row_index, column=j).value = remain_time
            sheet.cell(row=row_index, column=j).number_format = '#,##0.00'

def test_assignee(assignee, issue_key):
    issue_id = dic_story_map[issue_key][issue_id_col]
    is_assignee = None
    if dic_parent_map != {}:
        parent_id = dic_parent_map[issue_key]
        for item in dic_assignee[assignee]:
            if item == issue_id:
                is_assignee = True
            elif item in parent_id:
                is_assignee = True
    else:
        for item in dic_assignee[assignee]:
            if item == issue_id:
                is_assignee = True
    return is_assignee

def remaining_time(assignee, issue_key):
    remain_time = 0
    issue_id = dic_story_map[issue_key][issue_id_col]
    if issue_id in dic_assignee[assignee]:
        if dic_story_map[issue_key][remaining_col] != '':
            remain_time = remain_time + float(dic_story_map[issue_key][remaining_col])/3600
    else:
        for item in dic_parent_map[issue_key]:
            if item in dic_assignee[assignee]:
                key = dic_id_map[item]
                if dic_story_map[issue_key][remaining_col] != '':
                    remain_time = remain_time + float(dic_story_map[key][remaining_col])/3600
    return remain_time

# file name JIRA.csv, separators ,
with open('JIRA.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    data = list(readCSV)
    row_id = 1
    #maps each header to a column number based on the label
    for row in data:
        if row_id == 1:
            header = row
            issue_type_col  = map_columns(row, 'Issue Type')
            issue_key_col = map_columns(row, 'Issue key')
            issue_id_col = map_columns(row, 'Issue id')
            parent_id_col = map_columns(row,'Parent id')
            summary_col = map_columns(row,'Summary')
            assignee_col = map_columns(row, 'Assignee')
            remaining_col = map_columns(row, 'Î£ Remaining Estimate')
            row_id += 1
        else:
            #associates each issue id to a key
            dic_id_map[row[issue_id_col]] = row[issue_key_col]
            if row[assignee_col] not in dic_assignee.keys():
                dic_assignee[row[assignee_col]] = []
            dic_assignee[row[assignee_col]].append(row[issue_id_col])

    #build a list of the stories
    for row in data:
        if row == header:
            continue
        dic_story_map[row[issue_key_col]] = row
        if parent_id_col != None:
            list_stories(row[issue_key_col], row[issue_id_col], row[parent_id_col])
        row_id += 1

#generates the output file
output_file = load_workbook('Sprint_load.xlsx')
sheet = output_file.active
row_max = sheet.max_row
col_max = sheet.max_column

#checks for existing lines and updates them (specifically useful for update)
for key in dic_story_map:
    c = 13
    type = dic_story_map[key][issue_type_col]
    if type != 'Sub-task':
        if row_max >= 13:
            for i in range(14, row_max + 2):
                c = i
                issue_key = sheet.cell(row=i, column=1).value
                if issue_key != None:
                    if issue_key == key:
                        write_assignee(i, issue_key)
                        sheet.cell(row=i, column=col_max - 4).value = dic_story_map[issue_key][summary_col]
                        break
                    else:
                        continue
            if c == row_max + 1:
                if key not in new_stories:
                    new_stories.append(key)
        else:
            if key not in new_stories:
                new_stories.append(key)

#adds new lines in the end of the file (specifically useful for first creation)
index = row_max + 1
for item in new_stories:
    write_assignee(index, item)
    sheet.cell(row=index, column=1).value = item
    sheet.cell(row=index, column=col_max - 4).value = dic_story_map[item][summary_col]
    index += 1

#computes the sum of remaining assigned time
row_max = sheet.max_row
for j in range(2, col_max - 5):
    total = sheet.cell(row=12, column=j)
    top_cell_row = str(sheet.cell(row=15, column=j).row)
    top_cell_col = sheet.cell(row=15, column=j).column
    top_cell_col = str(ascii_uppercase[top_cell_col - 1])
    total_cell_row = str(sheet.cell(row=row_max, column=j).row)
    total_cell_col = sheet.cell(row=row_max, column=j).column
    total_cell_col = str(ascii_uppercase[total_cell_col - 1])
    total.value = '= SUM(' + top_cell_col + top_cell_row + ':' + total_cell_col + total_cell_row + ')'

#highlight the removed stories
cell_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid")
for i in range(14, row_max):
    story_key = sheet.cell(row=i, column=1)
    if story_key.value != None:
        if story_key.value not in dic_story_map.keys():
            for j in range(1, col_max - 5):
                sheet.cell(row=i, column=j).fill = cell_fill

output_file.save('Sprint_load_update.xlsx')